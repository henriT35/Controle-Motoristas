from io import BytesIO
from time import perf_counter
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Count, Q

from apps.core.models import SystemSettings
from apps.core.services import calculate_driver_metrics, completed_cte_ids, operational_date_map, operational_movements_for_period, parse_period
from apps.clients.models import Client
from apps.operations.models import CTe, DeliveryOccurrence
from apps.proofs.models import RetainedProof
from apps.ssw.models import ImportRun
from .models import GeneratedReport

REPORTS={
    "drivers":("Desempenho dos Motoristas","motoristas"),"proofs":("Comprovantes Retidos","comprovantes"),
    "clients":("Clientes","clientes"),"daily":("Operação Diária","operacao"),"ssw":("Importações SSW","ssw"),"finance":("Relatório Financeiro","financeiro"),
}

def _dataset(kind, start, end):
    if kind == "drivers":
        return [
            [m.driver.name, m.attempts, m.delivered, float(m.success_rate), m.clean_deliveries, float(m.clean_delivery_rate),
             m.retained, float(m.retention_rate), m.time_window_failures, float(m.time_window_rate), m.active_proofs,
             m.recovered_proofs, float(m.performance_score), float(m.ranking_score), float(m.productivity_score),
             m.sample_confidence, float(m.confidence_factor)]
            for m in calculate_driver_metrics(start, end)
        ], ["Motorista", "Tentativas", "Entregas", "Sucesso %", "Entregas limpas", "Entrega limpa %",
            "Retenções", "Retenção %", "Horário", "Horário %", "Comprovantes ativos",
            "Comprovantes resgatados", "Qualidade (simulação)", "Nota ajustada do ranking", "Produtividade", "Confiança", "Confiança estatística %"]

    if kind == "proofs":
        qs = RetainedProof.objects.filter(retained_at__date__range=(start, end)).exclude(
            status=RetainedProof.Status.CANCELED
        ).select_related("cte", "client", "address", "original_driver", "recovery_driver", "original_manifest")
        return [
            [p.cte.ctrc, p.invoice_number, p.client.name, p.client.cnpj,
             p.address.street if p.address else "", p.address.district if p.address else "",
             p.address.city if p.address else "", p.address.postal_code if p.address else "",
             p.original_driver.name, p.original_manifest.number if p.original_manifest else "",
             p.retained_at.date().isoformat(), p.days_retained, p.get_status_display(),
             p.recovery_driver.name if p.recovery_driver else "",
             p.recovered_at.date().isoformat() if p.recovered_at else "", float(p.freight_value),
             float(p.merchandise_value), float(p.weight_kg), p.volumes,
             "SIM" if p.client.proof_required_for_payment else "NÃO"]
            for p in qs
        ], ["CTRC", "NF", "Cliente", "CNPJ", "Endereço", "Bairro", "Município", "CEP",
            "Motorista da retenção", "Romaneio", "Retenção", "Dias retido", "Status",
            "Motorista recuperador", "Recuperado em", "Frete", "Valor mercadoria", "Peso kg", "Volumes",
            "Pagamento depende do comprovante"]

    if kind == "ssw":
        qs = ImportRun.objects.filter(created_at__date__range=(start, end))
        return [
            [r.created_at.strftime("%d/%m/%Y %H:%M"), r.get_kind_display(),
             f"{r.start_date:%d/%m/%Y}—{r.end_date:%d/%m/%Y}", r.new_count,
             r.updated_count, r.unchanged_count, r.get_status_display()]
            for r in qs
        ], ["Criado em", "Tipo", "Período", "Novos", "Atualizados", "Sem alteração", "Status"]

    movements = list(
        operational_movements_for_period(start, end)
        .filter(driver__is_test=False)
        .exclude(status__iexact="CANCELADO")
        .exclude(manifest__status__iexact="CANCELADO")
        .select_related("cte", "driver", "manifest", "client", "address")
    )
    route_dates = operational_date_map(start, end)
    delivered_ids = completed_cte_ids({m.cte_id for m in movements}, as_of=end)

    if kind == "clients":
        grouped = {}
        proof_counts = {
            row["client_id"]: row["count"]
            for row in RetainedProof.objects.filter(retained_at__date__range=(start, end))
            .exclude(status=RetainedProof.Status.CANCELED)
            .values("client_id").annotate(count=Count("id"))
        }
        for m in movements:
            if not m.client_id:
                continue
            row = grouped.setdefault(m.client_id, {"client": m.client, "ctes": set()})
            if m.cte_id in delivered_ids:
                row["ctes"].add(m.cte_id)
        rows = [
            [data["client"].name, data["client"].cnpj, len(data["ctes"]), proof_counts.get(client_id, 0),
             "SIM" if data["client"].proof_required_for_payment else "NÃO", data["client"].proof_payment_note]
            for client_id, data in sorted(grouped.items(), key=lambda item: item[1]["client"].name)
        ]
        return rows, ["Cliente", "CNPJ", "Entregas", "Retenções", "Pagamento depende do comprovante", "Observação da regra"]

    if kind == "daily":
        movement_ids = [m.pk for m in movements]
        flags = {mid: {"retention": False, "time": False} for mid in movement_ids}
        for occ in DeliveryOccurrence.objects.filter(movement_id__in=movement_ids, source="SSW_ROMANEIO").only("movement_id", "code", "description"):
            desc = (occ.description or "").upper()
            code = str(occ.code or "").strip()
            if code == "34" or "MERCADORIA EM CONFERENCIA NO CLIENTE" in desc:
                flags[occ.movement_id]["retention"] = True
            if code == "13" or "ENTREGA PREJUDICADA PELO HORARIO" in desc:
                flags[occ.movement_id]["time"] = True
        return [
            [(route_dates.get(m.manifest_id) or m.movement_date).isoformat(), m.driver.name,
             m.manifest.number, m.cte.ctrc, m.cte.invoice_number, m.client.name if m.client else "",
             m.address.street if m.address else "", m.address.district if m.address else "",
             m.address.city if m.address else "", m.attempt, "SIM" if m.cte_id in delivered_ids else "NÃO",
             "SIM" if flags.get(m.pk, {}).get("retention") else "NÃO",
             "SIM" if flags.get(m.pk, {}).get("time") else "NÃO", float(m.weight_kg), m.volumes,
             float(m.cte.freight_value), m.occurrence_text]
            for m in movements
        ], ["Data operacional", "Motorista", "Romaneio", "CTRC", "NF", "Cliente", "Endereço", "Bairro",
            "Município", "Tentativa", "Entregue", "Retenção", "Horário", "Peso kg", "Volumes", "Frete", "Ocorrência"]

    # Financeiro: um CT-e por chave dentro das rotas operacionais do período.
    ctes = {}
    for m in movements:
        ctes[m.cte_id] = m.cte
    return [
        [c.ctrc, c.invoice_number, c.client.name if c.client else "",
         "SIM" if c.client and c.client.proof_required_for_payment else "NÃO",
         float(c.freight_value), float(c.merchandise_value), float(c.weight_kg), c.current_status]
        for c in ctes.values()
    ], ["CTRC", "NF", "Cliente", "Pagamento depende do comprovante", "Frete", "Valor mercadoria", "Peso kg", "Status"]

@login_required
def index(request):
    start, end, label, mode = parse_period(request, SystemSettings.load().period_default)
    recent=GeneratedReport.objects.select_related("requested_by")[:20]
    totals = GeneratedReport.objects.aggregate(
        generated=Count("id"),
        exports=Count("id", filter=~Q(format=GeneratedReport.Format.HTML)),
    )
    period_query = request.GET.urlencode() or f"start={start.isoformat()}&end={end.isoformat()}"
    return render(request,"reports/index.html",{
        "report_cards":[(k,*v) for k,v in REPORTS.items()],"recent":recent,
        "generated_count":totals["generated"] or 0,"exports_count":totals["exports"] or 0,
        "period_start": start, "period_end": end, "period_label": label, "period_mode": mode,
        "period_query": period_query,
    })

@login_required
def preview(request,kind):
    if kind not in REPORTS: return HttpResponse("Tipo inválido",status=404)
    start,end,label,mode=parse_period(request,SystemSettings.load().period_default)
    rows,headers=_dataset(kind,start,end)
    GeneratedReport.objects.create(report_type=kind,start_date=start,end_date=end,format=GeneratedReport.Format.HTML,requested_by=request.user,file_name="visualização",row_count=len(rows))
    return render(request,"reports/preview.html",{"kind":kind,"title":REPORTS[kind][0],"rows":rows[:500],"headers":headers,"period_label":label,"period_start":start,"period_end":end,"period_mode":mode,"period_query":request.GET.urlencode()})

@login_required
def excel(request,kind):
    if kind not in REPORTS: return HttpResponse("Tipo inválido",status=404)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    start,end,label,mode=parse_period(request,SystemSettings.load().period_default); t0=perf_counter(); rows,headers=_dataset(kind,start,end)
    wb=Workbook(); ws=wb.active; ws.title="Relatório"; ws.append([REPORTS[kind][0]]); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(1,len(headers))); ws["A1"].font=Font(bold=True,size=16,color="FFFFFF"); ws["A1"].fill=PatternFill("solid",fgColor="0F172A"); ws["A1"].alignment=Alignment(horizontal="left")
    ws.append([f"Período: {label}"]); ws.append([]); ws.append(headers)
    for cell in ws[4]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="1D4ED8")
    for row in rows: ws.append(row)
    ws.freeze_panes="A5"
    last_col = get_column_letter(max(1, len(headers)))
    ws.auto_filter.ref = f"A4:{last_col}{max(ws.max_row, 4)}"
    for idx, col in enumerate(ws.iter_cols(min_col=1, max_col=max(1, len(headers))), start=1):
        letter=get_column_letter(idx)
        ws.column_dimensions[letter].width=min(45,max(12,max(len(str(c.value or "")) for c in col)+2))
        header = str(headers[idx-1]).lower() if idx-1 < len(headers) else ""
        if any(token in header for token in ("frete", "valor")):
            for cell in list(col)[4:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format='R$ #,##0.00'
        elif "peso" in header:
            for cell in list(col)[4:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format='#,##0.00'
    stream=BytesIO(); wb.save(stream); name=f"{REPORTS[kind][1]}_{start}_{end}.xlsx"; duration=int((perf_counter()-t0)*1000)
    GeneratedReport.objects.create(report_type=kind,start_date=start,end_date=end,format=GeneratedReport.Format.XLSX,requested_by=request.user,file_name=name,duration_ms=duration,row_count=len(rows),file_size=stream.tell())
    response=HttpResponse(stream.getvalue(),content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"); response["Content-Disposition"]=f'attachment; filename="{name}"'; return response

@login_required
def pdf(request,kind):
    if kind not in REPORTS: return HttpResponse("Tipo inválido",status=404)
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse("Dependência reportlab não instalada. Execute o instalador local novamente.",status=503)
    start,end,label,mode=parse_period(request,SystemSettings.load().period_default); t0=perf_counter(); rows,headers=_dataset(kind,start,end)
    stream=BytesIO(); doc=SimpleDocTemplate(stream,pagesize=landscape(A4),rightMargin=24,leftMargin=24,topMargin=24,bottomMargin=24); styles=getSampleStyleSheet(); story=[Paragraph(REPORTS[kind][0],styles["Title"]),Paragraph(f"Período: {label}",styles["Normal"]),Spacer(1,12)]
    data=[headers]+[[str(v) for v in row] for row in rows[:400]]; table=Table(data,repeatRows=1); table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0F172A")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),0.25,colors.HexColor("#CBD5E1")),("FONTSIZE",(0,0),(-1,-1),7),("VALIGN",(0,0),(-1,-1),"TOP"),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F8FAFC")])]))
    story.append(table); doc.build(story); name=f"{REPORTS[kind][1]}_{start}_{end}.pdf"; duration=int((perf_counter()-t0)*1000)
    GeneratedReport.objects.create(report_type=kind,start_date=start,end_date=end,format=GeneratedReport.Format.PDF,requested_by=request.user,file_name=name,duration_ms=duration,row_count=min(len(rows),400),file_size=stream.tell())
    response=HttpResponse(stream.getvalue(),content_type="application/pdf"); response["Content-Disposition"]=f'attachment; filename="{name}"'; return response
